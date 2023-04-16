from openpyxl import load_workbook
import datetime
import random

from .models import Analysis, AnalysisSumByDate
from .constants import DATA_FIELDS


def recalculate_sum_by_date():
    """Пересчитывает сумму сгруппированную по датам"""

    analyzes = Analysis.objects.all()
    analysis_sum_by_date = {}
    for analysis in analyzes:
        date = analysis.date
        if date not in analysis_sum_by_date:
            analysis_sum_by_date[date] = {}
            for field_name in DATA_FIELDS:
                analysis_sum_by_date[date][field_name] = 0
        for field_name in DATA_FIELDS:
            analysis_sum_by_date[date][field_name] += getattr(analysis, field_name)
    for date in analysis_sum_by_date:
        AnalysisSumByDate.objects.update_or_create(
            date=date, defaults=analysis_sum_by_date[date]
        )


def create_data_from_exel_file():
    """Загружает данные из эксель файла"""

    wb = load_workbook("data_analysis/init_data.xlsx")
    ws = wb.active
    analysis_data = {}
    for row in ws.iter_rows(min_row=4):
        analysis_data["pk"] = int(row[0].value)
        analysis_data["company_name"] = row[1].value

        for i, field_name in enumerate(DATA_FIELDS):
            analysis_data[field_name] = int(row[i + 2].value)
        analysis_data["date"] = datetime.date(
            2023, 7, random.randint(12, 16)
        )  # генерация дат случайна
        Analysis.objects.update_or_create(
            pk=analysis_data["pk"], defaults=analysis_data
        )
    recalculate_sum_by_date()
